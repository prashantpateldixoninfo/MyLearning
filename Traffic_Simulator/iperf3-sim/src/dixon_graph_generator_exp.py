import os
import sys
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

try:
    import xlwings as xw

    excel_available = True
except ImportError:
    excel_available = False


# Function to create a blank Excel file and save it in the background
def create_blank_excel(file_path):
    """Create a blank Excel file. Uses xlwings on Windows, openpyxl on Linux."""
    if sys.platform == "win32" and excel_available:
        # Use xlwings on Windows
        app = xw.App(visible=False)
        try:
            workbook = app.books.add()
            workbook.save(file_path)
            workbook.close()
        finally:
            app.quit()
    else:
        # Use openpyxl on Linux
        wb = Workbook()
        wb.save(file_path)

    print(f"Blank Excel file created at: {file_path}")


# Function to populate the Excel file with flexible data inputs from a dictionary
def populate_data(file_path, data_dict):
    if sys.platform == "win32" and excel_available:
        app = xw.App(visible=False)
        try:
            workbook = app.books.open(file_path)
            sheet = workbook.sheets[data_dict["sheet_num"]]
            sheet.name = data_dict["sheet_name"]

            # Insert headers and data based on dictionary keys
            headers = [
                data_dict["x_axis"],
                data_dict["first_header"],
                data_dict["sec_header"],
            ]
            data = [
                data_dict["x_axis_value"],
                data_dict["first_header_value"],
                data_dict["sec_header_value"],
            ]

            for col, (header, values) in enumerate(zip(headers, data), start=1):
                header_cell = sheet.range((1, col))
                data_range = sheet.range((2, col), (len(values) + 1, col))

                # Write header and data
                header_cell.value = header
                data_range.value = [[v] for v in values]

                # Format header: blue color text on white background, bold
                header_cell.api.Font.Color = 0xFF0000  # Blue text
                header_cell.api.Font.Bold = True

            # Apply grid lines around the entire data range
            data_range = sheet.range(
                (1, 1), (len(data_dict["x_axis_value"]) + 1, len(headers))
            )
            data_range.api.Borders.LineStyle = 1

            # Auto-fit columns
            sheet.autofit()

            # Save and close the workbook
            workbook.save()
            workbook.close()
        finally:
            app.quit()
    else:
        # Create a new workbook and select the active sheet
        wb = Workbook()
        sheet = wb.active
        sheet.title = data_dict["sheet_name"]

        # Insert headers
        headers = [
            data_dict["x_axis"],
            data_dict["first_header"],
            data_dict["sec_header"],
        ]
        data = [
            data_dict["x_axis_value"],
            data_dict["first_header_value"],
            data_dict["sec_header_value"],
        ]

        # Write headers to the first row
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            # Apply header style (blue text, bold)
            cell.font = Font(color="0000FF", bold=True)
            cell.fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )

        # Write data starting from the second row
        for col, (header, values) in enumerate(zip(headers, data), start=1):
            for row, value in enumerate(values, start=2):
                sheet.cell(row=row, column=col, value=value)

        # Apply gridlines (openpyxl automatically adds borders to cells)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=len(data_dict["x_axis_value"]) + 1,
            min_col=1,
            max_col=len(headers),
        ):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style="thin"),
                    right=openpyxl.styles.Side(style="thin"),
                    top=openpyxl.styles.Side(style="thin"),
                    bottom=openpyxl.styles.Side(style="thin"),
                )

        # Auto-size columns (openpyxl doesn't have auto-size, so we estimate width)
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in sheet.iter_rows(
                min_row=1,
                max_row=len(data_dict["x_axis_value"]) + 1,
                min_col=col,
                max_col=col,
            ):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

        # Save the workbook to the file
        wb.save(file_path)


# Function to add a chart to the Excel file and display it to the user
def add_chart(file_path, sheet_num, chart_title):
    if sys.platform == "win32" and excel_available:
        # xlwings implementation
        app = xw.App(visible=False)
        try:
            workbook = app.books.open(file_path)
            sheet = workbook.sheets[sheet_num]

            # Add a 2-D Line Chart (Lines with Markers) for the data
            chart = sheet.charts.add()
            chart.chart_type = "line_markers"
            chart.set_source_data(sheet.range("B1:C11"))
            chart.name = chart_title
            chart.api[1].HasTitle = True
            chart.api[1].ChartTitle.Text = chart_title
            chart.api[1].ApplyLayout(2)

            # Set x-axis to use manual labels from column A
            chart.api[1].Axes(1).CategoryNames = sheet.range("A2:A11").value

            # Position chart from E2 to L11
            chart.top = sheet.range("E2").top
            chart.left = sheet.range("E2").left
            chart.width = sheet.range("E2:L2").width
            chart.height = sheet.range("E2:E11").height

            # Save and close the workbook
            workbook.save()
            workbook.close()
        finally:
            app.quit()
    else:
        # openpyxl implementation
        wb = load_workbook(file_path)
        sheet = wb.worksheets[sheet_num]

        # Create a LineChart
        chart = LineChart()
        chart.title = chart_title

        # Set data for the chart
        data = Reference(sheet, min_col=2, min_row=1, max_col=3, max_row=11)
        chart.add_data(data, titles_from_data=True)

        # Set categories (x-axis labels)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=11)
        chart.set_categories(categories)

        # Position chart from E2 to L11
        chart.anchor = "E2"

        # Add the chart to the sheet
        sheet.add_chart(chart, "E2")

        # Save the workbook
        wb.save(file_path)


def display_chart(file_path):
    app = xw.App(visible=True)
    workbook = app.books.open(file_path)
    sheet = workbook.sheets[0]  # Assuming we work with the first sheet


def main():
    # Define file path and data dictionary
    file_path = os.path.join(os.getcwd(), "line_chart_example_final.xlsx")
    data_dict = {
        "sheet_num": 0,  # First sheet in the workbook
        "sheet_name": "Bitrate",
        "x_axis": "Bitrate",
        "first_header": "Zyxel-2.4_Bitrate",
        "sec_header": "DXB-2.4_Bitrate",
        "x_axis_value": [50, 80, 120, 200, 300, 400, 500, 700, 800, 1000],
        "first_header_value": [0, 6.5, 23, 22, 29, 53, 40, 57, 43, 45],
        "sec_header_value": [0, 5.4, 13, 32, 39, 42, 42, 51, 48, 51],
    }
    chart_title = "Bitrate Comparison"

    # Execute each function sequentially
    create_blank_excel(file_path)
    populate_data(file_path, data_dict)
    add_chart(file_path, 0, chart_title)
    display_chart(file_path)


if __name__ == "__main__":
    main()
