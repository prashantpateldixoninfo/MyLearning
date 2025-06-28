import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import Font
from copy import copy
from openpyxl.cell import MergedCell

# Input to Output column index mapping (0-based)
COLUMN_INDEX_MAP = {
    0: 0,   # A -> A
    1: 1,   # B -> B
    3: 2,   # D -> C
    4: 3,   # E -> D
    5: 4,   # F -> E
    7: 5,   # H -> F
    8: 6,   # I -> G
    9: 7,   # J -> H
    10: 8,  # K -> I
    11: 9,  # L -> J
    12: 10, # M -> K
    13: 11, # N -> L
    14: 17, # O -> R
    15: 12, # P -> M
    16: 13, # Q -> N
    17: 14, # R -> O
    18: 15, # S -> P
    19: 16, # T -> Q
}

SHEET_NAMES = ["Stores & IQC", "SMT", "MI", "BLT & FAT", "PACKING"]

def copy_range_formatting(src_ws, tgt_ws, cell_range):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    for merged_range in src_ws.merged_cells.ranges:
        minc, minr, maxc, maxr = range_boundaries(str(merged_range))
        if min_row <= minr <= max_row and min_col <= minc <= max_col:
            tgt_ws.merge_cells(start_row=minr, start_column=minc, end_row=maxr, end_column=maxc)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            src_cell = src_ws.cell(row=row, column=col)
            tgt_cell = tgt_ws.cell(row=row, column=col, value=src_cell.value)
            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.alignment = copy(src_cell.alignment)
                tgt_cell.number_format = src_cell.number_format

def copy_row_formatting(src_ws, tgt_ws, src_row_idx, tgt_row_idx):
    for col in range(1, tgt_ws.max_column + 1):
        src_cell = src_ws.cell(row=src_row_idx, column=col)
        tgt_cell = tgt_ws.cell(row=tgt_row_idx, column=col, value=src_cell.value)
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.alignment = copy(src_cell.alignment)
            tgt_cell.number_format = src_cell.number_format

def transform_excel(input_file, output_file):
    wb_input = load_workbook(input_file)
    wb_output = load_workbook(output_file)

    if "PFMEA-Format" not in wb_output.sheetnames:
        print("[ERROR] 'PFMEA-Format' sheet is missing in output/template file")
        sys.exit(1)

    template_ws = wb_output["PFMEA-Format"]

    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb_input.sheetnames:
            print(f"[SKIP] Sheet '{sheet_name}' not found in input file")
            continue

        input_ws = wb_input[sheet_name]

        if sheet_name in wb_output.sheetnames:
            del wb_output[sheet_name]
        new_ws = wb_output.create_sheet(title=sheet_name)

        copy_range_formatting(template_ws, new_ws, "A1:T8")

        header_row = input_ws[9]  # Row 10 (1-based)
        for orig_idx, new_idx in COLUMN_INDEX_MAP.items():
            if orig_idx < len(header_row):
                cell_value = header_row[orig_idx].value
                if cell_value:
                    target_cell = new_ws.cell(row=8, column=new_idx + 1)
                    if not isinstance(target_cell, MergedCell):
                        target_cell.value = cell_value
                        target_cell.font = Font(bold=True)

        last_row_written = 8
        for row_idx, row in enumerate(input_ws.iter_rows(min_row=10), start=9):
            for orig_idx, new_idx in COLUMN_INDEX_MAP.items():
                if orig_idx < len(row):
                    cell = new_ws.cell(row=row_idx, column=new_idx + 1)
                    source_cell = row[orig_idx]
                    cell.value = source_cell.value
                    if source_cell.has_style:
                        cell.font = copy(source_cell.font)
                        cell.fill = copy(source_cell.fill)
                        cell.alignment = copy(source_cell.alignment)
                        cell.border = copy(source_cell.border)

            e = new_ws.cell(row=row_idx, column=5).coordinate
            g = new_ws.cell(row=row_idx, column=7).coordinate
            j = new_ws.cell(row=row_idx, column=10).coordinate
            n = new_ws.cell(row=row_idx, column=14).coordinate
            o = new_ws.cell(row=row_idx, column=15).coordinate
            p = new_ws.cell(row=row_idx, column=16).coordinate

            new_ws.cell(row=row_idx, column=11, value=f"={e}*{g}*{j}")
            new_ws.cell(row=row_idx, column=17, value=f"={n}*{o}*{p}")
            last_row_written = row_idx

        # ✅ Overwrite last row with row 41 from 'PFMEA-Format'
        copy_row_formatting(template_ws, new_ws, 41, last_row_written)

    wb_output.save(output_file)
    print(f"[OK] Output saved as: {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python excel_transformer.py <input.xlsx> <output_template.xlsx>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    if not os.path.isfile(input_file):
        print(f"[ERROR] File not found: {input_file}")
        sys.exit(1)

    transform_excel(input_file, output_file)
