import os
import glob
import re
import sys
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font


def extract_serial(lines):
    for line in lines:
        if line.startswith("UADCSTART|"):
            parts = line.strip().split("|")
            if len(parts) > 1:
                return parts[1]
    return None


def extract_test_items(lines):
    pattern = r'^PRODUCTSTEP\|TestItem\|([^|]+)\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|([^|]*)'
    data = {}
    for line in lines:
        match = re.match(pattern, line)
        if match:
            test_name = match.group(1).strip()
            value_str = match.group(2).strip()
            try:
                value = float(value_str)
                data[test_name] = value
            except ValueError:
                continue
    return data


def extract_min_max_per_test(lines):
    result = {}
    for line in lines:
        if line.startswith("PRODUCTSTEP|TestItem|"):
            parts = line.strip().split("|")
            if len(parts) >= 11:
                test_key = parts[2].strip()
                try:
                    min_val = float(parts[8])
                    max_val = float(parts[9])
                    result[test_key] = (min_val, max_val)
                except ValueError:
                    continue
    return result


def process_folder(base_folder, master_data, master_min_max):
    serial_number_folder = os.path.basename(base_folder)
    log_files = sorted(glob.glob(os.path.join(base_folder, "*.dat")))
    if not log_files:
        print(f"[SKIP] No .dat files in {base_folder}")
        return

    serial_numbers = []
    all_data = {}
    file_names = []

    for file_path in log_files:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        serial = extract_serial(lines)
        serial_numbers.append(serial)

        if serial != serial_number_folder:
            raise ValueError(f"[ERROR] Serial mismatch in {file_path}: found {serial}, expected {serial_number_folder}")

        file_names.append(os.path.basename(file_path))
        test_data = extract_test_items(lines)

        for test, value in test_data.items():
            if test not in all_data:
                all_data[test] = []
            all_data[test].append(value)

            if test not in master_data:
                master_data[test] = {}
            if serial not in master_data[test]:
                master_data[test][serial] = []
            master_data[test][serial].append(value)

    with open(log_files[-1], "r", encoding="utf-8") as f:
        last_lines = f.readlines()
    min_max_data = extract_min_max_per_test(last_lines)

    for test, (min_val, max_val) in min_max_data.items():
        if test not in master_min_max:
            master_min_max[test] = (min_val, max_val)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    bold = Font(bold=True)

    ws["A1"] = serial_number_folder
    ws["A1"].font = bold
    for idx, file_name in enumerate(file_names):
        col = chr(66 + idx)
        ws[f"{col}1"] = file_name
        ws[f"{col}1"].font = bold

    avg_col = chr(66 + len(file_names))
    min_col = chr(66 + len(file_names) + 1)
    max_col = chr(66 + len(file_names) + 2)

    ws[f"{avg_col}1"] = "Average"
    ws[f"{avg_col}1"].font = bold
    ws[f"{min_col}1"] = "Min"
    ws[f"{min_col}1"].font = bold
    ws[f"{max_col}1"] = "Max"
    ws[f"{max_col}1"].font = bold

    for row_idx, (test_name, values) in enumerate(all_data.items(), start=2):
        ws[f"A{row_idx}"] = test_name
        for col_offset, value in enumerate(values):
            col = chr(66 + col_offset)
            ws[f"{col}{row_idx}"] = value
        if values:
            ws[f"{avg_col}{row_idx}"] = round(mean(values), 3)
        if test_name in min_max_data:
            min_val, max_val = min_max_data[test_name]
            ws[f"{min_col}{row_idx}"] = min_val
            ws[f"{max_col}{row_idx}"] = max_val

    output_file = os.path.join(base_folder, f"{serial_number_folder}_Test_Result.xlsx")
    wb.save(output_file)
    print(f"[OK] Excel saved: {output_file}")


def generate_master_excel(all_results, master_min_max, base_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Results"
    bold = Font(bold=True)

    ws["A1"] = "Test Item"
    ws["A1"].font = bold
    serials = sorted({serial for data in all_results.values() for serial in data})

    for idx, serial in enumerate(serials):
        col = chr(66 + idx)
        ws[f"{col}1"] = serial
        ws[f"{col}1"].font = bold

    min_col = chr(66 + len(serials))
    max_col = chr(66 + len(serials) + 1)
    ws[f"{min_col}1"] = "Min"
    ws[f"{min_col}1"].font = bold
    ws[f"{max_col}1"] = "Max"
    ws[f"{max_col}1"].font = bold

    for row_idx, (test_name, serial_data) in enumerate(all_results.items(), start=2):
        ws[f"A{row_idx}"] = test_name
        for col_idx, serial in enumerate(serials):
            col = chr(66 + col_idx)
            val = serial_data.get(serial, [])
            if isinstance(val, list):
                ws[f"{col}{row_idx}"] = round(mean(val), 3) if val else ""
            else:
                ws[f"{col}{row_idx}"] = val

        if test_name in master_min_max:
            min_val, max_val = master_min_max[test_name]
            ws[f"{min_col}{row_idx}"] = min_val
            ws[f"{max_col}{row_idx}"] = max_val

    output_file = os.path.join(base_dir, "Master_Test_Result.xlsx")
    wb.save(output_file)
    print(f"[OK] Master Excel saved: {output_file}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python regex_all_folders.py \"<parent_folder_path>\"")
        sys.exit(1)

    input_path = sys.argv[1]
    base_dir = os.path.abspath(input_path)

    if not os.path.isdir(base_dir):
        print(f"[ERROR] Invalid directory: {base_dir}")
        sys.exit(1)

    subfolders = [f.path for f in os.scandir(base_dir) if f.is_dir()]
    if not subfolders:
        print("[INFO] No subdirectories found.")
        return

    all_results = {}
    master_min_max = {}
    for folder in subfolders:
        try:
            process_folder(folder, all_results, master_min_max)
        except Exception as e:
            print(f"[FAIL] Error processing {folder}: {e}")

    generate_master_excel(all_results, master_min_max, base_dir)


if __name__ == "__main__":
    main()
