import sys
import re
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font

def process_file(input_txt):
    serial_number = None
    test_items = []

    with open(input_txt, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    # Extract serial number
    for line in lines:
        if line.startswith("UADCSTART|"):
            parts = line.strip().split('|')
            if len(parts) > 1:
                serial_number = parts[1]
            break

    # Extract TestItem with value, min, max
    pattern = r'^PRODUCTSTEP\|TestItem\|([^|]+)\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|([^|]*)\|([^|]*)\|([^|]*)'
    for line in lines:
        match = re.match(pattern, line)
        if match:
            test_name = match.group(1).strip()
            value = match.group(2).strip()
            min_val = match.group(3).strip()
            max_val = match.group(4).strip()
            test_items.append((test_name, value, min_val, max_val))

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    bold_font = Font(bold=True)
    ws['A1'] = f"Serial Number: {serial_number}"
    ws['A1'].font = bold_font
    ws['B1'], ws['C1'], ws['D1'] = "Value", "Min", "Max"
    ws['B1'].font = ws['C1'].font = ws['D1'].font = bold_font

    for idx, (test_name, value, min_val, max_val) in enumerate(test_items, start=2):
        ws[f"A{idx}"] = test_name
        ws[f"B{idx}"] = value
        ws[f"C{idx}"] = min_val
        ws[f"D{idx}"] = max_val

    # Save Excel file next to the input file
    base_name = os.path.splitext(os.path.basename(input_txt))[0]
    folder = os.path.dirname(input_txt)
    output_path = os.path.abspath(os.path.join(folder, f"{base_name}.xlsx"))
    wb.save(output_path)
    print(f"✔ Excel saved: {output_path}")

def main():
    if len(sys.argv) != 2:
        print("Usage: python Regex.py \"Wifi ports3\\*\"")
        sys.exit(1)

    top_folder = sys.argv[1]

    # Match all .dat files in one-level subdirectories
    pattern = os.path.join(top_folder, "*", "*.dat")
    matched_files = glob.glob(pattern)

    if not matched_files:
        print(f"No .dat files found in: {pattern}")
        sys.exit(1)

    for file_path in matched_files:
        process_file(file_path)

if __name__ == "__main__":
    main()
